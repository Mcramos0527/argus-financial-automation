# app/services/exporter.py
# Genera los archivos Excel de salida de ARGUS.

import logging
from datetime import date, datetime
from pathlib import Path
from typing import List

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from app.models import BankSummary, CajaEntry, Transaction

logger = logging.getLogger("argus.exporter")

# ── Estilos ───────────────────────────────────────────────────────────────────

def _header_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

FILL_HEADER_BLUE   = _header_fill("1F4E79")
FILL_HEADER_GRAY   = _header_fill("404040")
FILL_DD_SRL        = _header_fill("BDD7EE")   # azul claro DD SRL
FILL_D_CIA         = _header_fill("FFEB9C")   # amarillo D y CIA
FILL_COBRO         = _header_fill("E2EFDA")   # verde claro cobro
FILL_PAGO          = _header_fill("FCE4D6")   # rojo claro pago
FILL_INTERNO       = _header_fill("EDEDED")   # gris interno

FONT_WHITE_BOLD = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
FONT_BOLD       = Font(bold=True, name="Calibri", size=10)
FONT_NORMAL     = Font(name="Calibri", size=10)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center")
ALIGN_RIGHT  = Alignment(horizontal="right",  vertical="center")

PESO_FORMAT  = '#,##0.00'
DATE_FORMAT  = 'DD/MM/YYYY'


def _set_header_row(ws, headers: list, fill: PatternFill, row: int = 1):
    """Escribe una fila de encabezados con estilo."""
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font      = FONT_WHITE_BOLD
        cell.fill      = fill
        cell.border    = THIN_BORDER
        cell.alignment = ALIGN_CENTER


def _auto_width(ws, min_w: int = 10, max_w: int = 40):
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


# ── Exportador principal ──────────────────────────────────────────────────────

class Exporter:

    def export_all(
        self,
        transactions: List[Transaction],
        summaries: List[BankSummary],
        caja_entries: List[CajaEntry],
        output_folder: str,
    ) -> List[str]:
        """
        Genera los 4 archivos Excel de salida.
        Retorna lista de rutas generadas.
        """
        folder = Path(output_folder)
        folder.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        generated = []

        # 1. Transacciones normalizadas
        path = folder / f"argus_movimientos_normalizados_{timestamp}.xlsx"
        self._export_transactions(transactions, path)
        generated.append(str(path))
        logger.info(f"Exportado: {path.name}")

        # 2. Resumen bancario diario
        path = folder / f"argus_resumen_bancario_{timestamp}.xlsx"
        self._export_summary(summaries, path)
        generated.append(str(path))
        logger.info(f"Exportado: {path.name}")

        # 3. Export para Caja Fábrica Digital
        path = folder / f"argus_export_caja_{timestamp}.xlsx"
        self._export_caja(caja_entries, path)
        generated.append(str(path))
        logger.info(f"Exportado: {path.name}")

        return generated

    # ── Archivo 1: Movimientos normalizados ───────────────────────────────────

    def _export_transactions(self, transactions: List[Transaction], path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Movimientos"

        # Cabecera ARGUS
        ws.merge_cells("A1:T1")
        title_cell = ws["A1"]
        title_cell.value     = "ARGUS — Movimientos Normalizados"
        title_cell.font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
        title_cell.fill      = _header_fill("1F3864")
        title_cell.alignment = ALIGN_CENTER
        ws.row_dimensions[1].height = 22

        headers = [
            "Pestaña", "Empresa", "Banco", "Fecha", "Fecha Valor",
            "Descripción", "Detalle", "Débito", "Crédito", "Importe Neto",
            "Saldo", "Nro Referencia", "Nro Cheque", "Cod Concepto",
            "Tipo Concepto", "Canal", "Sucursal",
            "Cat. Código", "Cat. Nombre", "Tipo Movimiento",
        ]
        _set_header_row(ws, headers, FILL_HEADER_BLUE, row=2)
        ws.row_dimensions[2].height = 16
        ws.freeze_panes = "A3"

        for row_idx, tx in enumerate(transactions, start=3):
            row_data = [
                tx.pestaña, tx.empresa, tx.banco,
                tx.fecha, tx.fecha_valor,
                tx.descripcion, tx.detalle,
                tx.debito or None, tx.credito or None, tx.importe_neto,
                tx.saldo,
                tx.nro_referencia, tx.nro_cheque, tx.cod_concepto,
                tx.tipo_concepto, tx.canal, tx.sucursal,
                tx.categoria_codigo, tx.categoria_nombre, tx.tipo_movimiento,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font   = FONT_NORMAL
                cell.border = THIN_BORDER

                # Formato numérico
                if col_idx in (8, 9, 10, 11):  # débito, crédito, neto, saldo
                    cell.number_format = PESO_FORMAT
                    cell.alignment     = ALIGN_RIGHT

                # Formato fecha
                if col_idx in (4, 5):
                    cell.number_format = DATE_FORMAT
                    cell.alignment     = ALIGN_CENTER

                # Color por tipo movimiento
                tipo = tx.tipo_movimiento
                if col_idx == 20:
                    if tipo == "COBRO":
                        cell.fill = FILL_COBRO
                        cell.font = Font(name="Calibri", size=10, bold=True, color="375623")
                    elif tipo == "PAGO":
                        cell.fill = FILL_PAGO
                        cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
                    elif tipo == "INTERNO":
                        cell.fill = FILL_INTERNO

                # Color por empresa en columna empresa
                if col_idx == 2:
                    if "Dario" in (tx.empresa or ""):
                        cell.fill = FILL_DD_SRL
                    elif "Cia" in (tx.empresa or "") or "cia" in (tx.empresa or ""):
                        cell.fill = FILL_D_CIA

        _auto_width(ws)
        wb.save(path)

    # ── Archivo 2: Resumen bancario ───────────────────────────────────────────

    def _export_summary(self, summaries: List[BankSummary], path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Bancos del Día"

        # Título
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value     = f"ARGUS — Resumen Bancario del Día  |  Generado: {date.today().strftime('%d/%m/%Y')}"
        c.font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
        c.fill      = _header_fill("1F3864")
        c.alignment = ALIGN_CENTER
        ws.row_dimensions[1].height = 22

        headers = [
            "Empresa", "Banco / Cuenta",
            "Saldo Actual", "Cobros del Día", "Pagos del Día",
            "Gastos Bancarios", "Intereses", "Movimientos",
        ]
        _set_header_row(ws, headers, FILL_HEADER_BLUE, row=2)
        ws.freeze_panes = "A3"

        current_company = None
        row_idx = 3

        for s in summaries:
            # Separador de empresa
            if s.empresa != current_company:
                current_company = s.empresa
                ws.merge_cells(f"A{row_idx}:H{row_idx}")
                c = ws.cell(row=row_idx, column=1, value=f"  {s.empresa}")
                c.font      = Font(bold=True, size=11, name="Calibri",
                                   color="FFFFFF")
                c.fill      = FILL_DD_SRL if "Dario" in s.empresa else FILL_D_CIA
                c.alignment = ALIGN_LEFT
                ws.row_dimensions[row_idx].height = 18
                row_idx += 1

            row_data = [
                s.empresa, s.banco,
                s.saldo_actual, s.cobros_dia, s.pagos_dia,
                s.gastos_dia, s.intereses_dia, s.movimientos_count,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font   = FONT_NORMAL
                cell.border = THIN_BORDER
                if col_idx in (3, 4, 5, 6, 7):
                    cell.number_format = PESO_FORMAT
                    cell.alignment     = ALIGN_RIGHT
                if col_idx == 8:
                    cell.alignment = ALIGN_CENTER

            row_idx += 1

        # Totales por empresa
        row_idx += 1
        ws.merge_cells(f"A{row_idx}:B{row_idx}")
        ws.cell(row=row_idx, column=1, value="TOTAL GENERAL").font = FONT_BOLD
        total_saldo  = sum(s.saldo_actual for s in summaries)
        total_cobros = sum(s.cobros_dia   for s in summaries)
        total_pagos  = sum(s.pagos_dia    for s in summaries)
        for col, val in zip([3, 4, 5], [total_saldo, total_cobros, total_pagos]):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font           = FONT_BOLD
            c.number_format  = PESO_FORMAT
            c.alignment      = ALIGN_RIGHT
            c.fill           = _header_fill("D6DCE4")

        _auto_width(ws)
        wb.save(path)

    # ── Archivo 3: Export Caja ─────────────────────────────────────────────────

    def _export_caja(self, entries: List[CajaEntry], path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Export Caja"

        ws.merge_cells("A1:I1")
        c = ws["A1"]
        c.value     = "ARGUS — Export para Caja Fábrica Digital"
        c.font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
        c.fill      = _header_fill("375623")
        c.alignment = ALIGN_CENTER
        ws.row_dimensions[1].height = 22

        headers = [
            "DIA", "Fecha", "NRO TIPO", "TIPO", "IMPORTE",
            "DESCRIPCIÓN", "Canal", "Empresa", "Banco",
        ]
        _set_header_row(ws, headers, _header_fill("375623"), row=2)
        ws.freeze_panes = "A3"

        for row_idx, entry in enumerate(entries, start=3):
            row_data = [
                entry.dia, entry.fecha,
                entry.nro_tipo, entry.tipo, entry.importe,
                entry.descripcion, entry.canal,
                entry.empresa, entry.banco,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font   = FONT_NORMAL
                cell.border = THIN_BORDER
                if col_idx == 5:
                    cell.number_format = PESO_FORMAT
                    cell.alignment     = ALIGN_RIGHT
                if col_idx == 2:
                    cell.number_format = DATE_FORMAT
                    cell.alignment     = ALIGN_CENTER
                if col_idx in (1, 3):
                    cell.alignment = ALIGN_CENTER

        _auto_width(ws)
        wb.save(path)
