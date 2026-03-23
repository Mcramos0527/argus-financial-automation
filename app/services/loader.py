# app/services/loader.py
# Servicio de carga de archivos Excel del cliente Delfabro.
# Lee los 3 archivos y los deja disponibles para el resto del pipeline.

import logging
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl import load_workbook

from app.config import SHEET_CONFIG, NON_BANK_SHEETS

logger = logging.getLogger("argus.loader")


class ExcelLoader:
    """Carga y valida los tres archivos Excel del proceso Delfabro."""

    def __init__(self):
        self.wb_movimientos: Optional[openpyxl.Workbook] = None
        self.wb_saldos: Optional[openpyxl.Workbook] = None
        self.wb_caja: Optional[openpyxl.Workbook] = None
        self._paths: Dict[str, Path] = {}

    # ── Carga ────────────────────────────────────────────────────────────────

    def load_movimientos(self, path: str) -> Tuple[bool, str]:
        """Carga el archivo de Movimientos Diarios Bancarios."""
        try:
            p = Path(path)
            if not p.exists():
                return False, f"Archivo no encontrado: {path}"
            self.wb_movimientos = load_workbook(p, read_only=True, data_only=True)
            self._paths["movimientos"] = p
            detected = self._detect_bank_sheets()
            logger.info(f"Movimientos cargado: {p.name} — {len(detected)} pestañas bancarias")
            return True, f"OK — {len(detected)} cuentas detectadas"
        except Exception as e:
            logger.error(f"Error cargando movimientos: {e}")
            return False, f"Error: {e}"

    def load_saldos(self, path: str) -> Tuple[bool, str]:
        """Carga el archivo de Saldos del Día."""
        try:
            p = Path(path)
            if not p.exists():
                return False, f"Archivo no encontrado: {path}"
            self.wb_saldos = load_workbook(p, read_only=True, data_only=True)
            self._paths["saldos"] = p
            logger.info(f"Saldos cargado: {p.name}")
            return True, "OK"
        except Exception as e:
            logger.error(f"Error cargando saldos: {e}")
            return False, f"Error: {e}"

    def load_caja(self, path: str) -> Tuple[bool, str]:
        """Carga el archivo de Caja Fábrica Digital."""
        try:
            p = Path(path)
            if not p.exists():
                return False, f"Archivo no encontrado: {path}"
            self.wb_caja = load_workbook(p, read_only=True, data_only=True)
            self._paths["caja"] = p
            logger.info(f"Caja cargado: {p.name}")
            return True, "OK"
        except Exception as e:
            logger.error(f"Error cargando caja: {e}")
            return False, f"Error: {e}"

    # ── Detección de pestañas ─────────────────────────────────────────────────

    def _detect_bank_sheets(self) -> List[str]:
        """Retorna las pestañas bancarias conocidas que existen en el workbook."""
        if not self.wb_movimientos:
            return []
        available = set(self.wb_movimientos.sheetnames)
        known = set(SHEET_CONFIG.keys())
        found = [s for s in self.wb_movimientos.sheetnames if s in known]
        unknown_new = available - known - NON_BANK_SHEETS
        if unknown_new:
            logger.warning(f"Pestañas desconocidas (no se procesarán): {unknown_new}")
        return found

    def get_bank_sheets(self) -> List[str]:
        """Lista de pestañas bancarias detectadas."""
        return self._detect_bank_sheets()

    # ── Lectura de filas ──────────────────────────────────────────────────────

    def get_sheet_rows(self, sheet_name: str) -> List[tuple]:
        """Devuelve todas las filas de una pestaña del workbook de movimientos."""
        if not self.wb_movimientos:
            return []
        if sheet_name not in self.wb_movimientos.sheetnames:
            logger.warning(f"Pestaña '{sheet_name}' no encontrada")
            return []
        ws = self.wb_movimientos[sheet_name]
        return list(ws.iter_rows(values_only=True))

    def get_categories(self) -> Dict[int, str]:
        """Lee la tabla de categorías contables y retorna {codigo: nombre}."""
        cats: Dict[int, str] = {}
        if not self.wb_movimientos:
            return cats
        sheet_name = "CATEGORIAS CONTABLES"
        if sheet_name not in self.wb_movimientos.sheetnames:
            logger.warning("Pestaña CATEGORIAS CONTABLES no encontrada")
            return cats
        ws = self.wb_movimientos[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                try:
                    code = int(row[0])
                    name = str(row[1]).strip()
                    cats[code] = name
                except (ValueError, TypeError):
                    continue
        logger.info(f"Categorías cargadas: {len(cats)}")
        return cats

    def get_saldos_rows(self) -> List[tuple]:
        """Devuelve las filas de la pestaña BANCOS DEL DIA."""
        if not self.wb_saldos:
            return []
        sheet_name = "BANCOS DEL DIA"
        if sheet_name not in self.wb_saldos.sheetnames:
            logger.warning("Pestaña BANCOS DEL DIA no encontrada")
            return []
        ws = self.wb_saldos[sheet_name]
        return list(ws.iter_rows(values_only=True))

    def get_caja_months(self) -> List[str]:
        """Devuelve las pestañas de meses disponibles en Caja."""
        if not self.wb_caja:
            return []
        return [s for s in self.wb_caja.sheetnames if s not in {"acumulado", "Categorías"}]

    def is_ready(self) -> bool:
        """True si los 3 archivos están cargados."""
        return all([self.wb_movimientos, self.wb_saldos, self.wb_caja])
